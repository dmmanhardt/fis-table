# -*- coding: utf-8 -*-
"""
Created on Mon Jul  2 16:33:05 2018

@author: MANHARDTD
"""

from openpyxl import load_workbook
from pandas import DataFrame
import os

#create tables to store sig wave height values to add to dataframe and output to excel sheet
fileStore = []
WaveHeightStore = []
WavePeriodStore = []

def FindTWEL(fileToOpen):
    #open file storing 1% TWEL info
    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        if filename.endswith('TWEL.xlsx'):
            TWELfile = filename
            os.getcwd()
            print(TWELfile)
            wb = load_workbook(filename = TWELfile, data_only = True)
            print("Loaded file")
            sheet_ranges = wb["sheet1"]
            for i in range(2, 200):
                checkFileName = sheet_ranges.cell(row = i, column = 1).value
                # checkFileName = checkFileName + ".xlsm"
                if checkFileName == fileToOpen:
                    TWEL = sheet_ranges.cell(row = i, column = 3).value
                    return TWEL
            
        

def FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn, event_twel_column):
    TWEL = FindTWEL(fileToOpen)
    #find the max of max TWELs from doc sheet
    sheet_ranges = wb['Doc']

    #change working sheet to EventSummary sheet
    sheet_ranges = wb["EventSummary"]
    
    #search MaxTWEL values in EventSummary to find corresponding row/hour
    sheet = wb.get_sheet_by_name("EventSummary")
    OriginalDifferenceBetweenTWELs = 100
    newTWEL = False
    for i in range(3, 152):
        valueCheck = sheet.cell(row = i, column = TWLColumn).value
        checkDifference = abs(valueCheck - TWEL)
        if checkDifference < OriginalDifferenceBetweenTWELs:
            OriginalDifferenceBetweenTWELs = checkDifference
            newTWELRow = i
            closest_TWEL = valueCheck
            newTWEL = True
            EventNumber = sheet.cell(row = newTWELRow, column = EventColumn).value           
    
    if newTWEL == True:
        #use event # to navigate to the correct event sheet
        EventNumberSheet = "Event" + str(EventNumber)
        sheet = wb.get_sheet_by_name(EventNumberSheet)
            
        #use MaxWaveHeight to find correct row in H0 column 
        for i in range(7, 1000):
            ValueToCheck = sheet.cell(row = i, column = event_twel_column).value
            if ValueToCheck == closest_TWEL:
                #use periodrow to find waveperiod
                wave_height = sheet.cell(row = i, column = EventWaveHeightColumn).value
                WavePeriod = sheet.cell(row = i, column = EventWavePeriodColumn).value
                break
            else:
                wave_height = "ERROR no wave height found"
                WavePeriod = "ERROR no wave period found"
                
            #write values to excel file
        fileStore.append(fileToOpen)
        WaveHeightStore.append(wave_height)
        WavePeriodStore.append(WavePeriod)


#Path where excel files are located, needs to be updated for each county
    
directory = input("What is the directory where the runup files are stored? ")
directory = os.chdir(directory)
#loop through each file in the directory and opens file
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith('.xlsm') and "~$" not in filename:
        fileToOpen = filename
        os.getcwd()
        print(fileToOpen)
        wb = load_workbook(filename = fileToOpen, data_only = True)
        print("Loaded file")

#check what method (type 1-3) and update parameters for that method, then run FindSigWaveValues()

        if "Type 1_Stockdon_Erosion" in fileToOpen:
            print("Opened file as type 1 eroded")
            DocCell = 'C25'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 19
            event_twel_column = 24
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7
            
            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn, event_twel_column)
        
        elif "Type 1" in fileToOpen:
            print("opened file as type 1")
            DocCell = 'C25'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 20
            event_twel_column = 24
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7

            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn, event_twel_column)
            
        #if type 2 method cells change
        elif "Type 2" in fileToOpen:
            print("opened file as type 2")
            DocCell = 'C33'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 20
            event_twel_column = 46
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7

            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn, event_twel_column)
        
        #if type 3 method, cells and sheets change
        elif "Type 3" in fileToOpen:
            print("opened file as type 3")
            DocCell = 'C24'
            TWLColumn = 5
            EventColumn = 1
            WaveHeightColumn = 18
            event_twel_column = 42
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7

            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn, event_twel_column)
        else:
            print("ERROR")
            quit()
    #write values to excel file
    df = DataFrame({'File': fileStore, 'WaveHeight': WaveHeightStore, 'WavePeriod': WavePeriodStore})
    print(df)
    df.to_excel('NewValues.xlsx', sheet_name = 'sheet1', index = False)