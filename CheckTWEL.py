# -*- coding: utf-8 -*-
"""
Created on Tue Jul  3 10:51:24 2018

@author: MANHARDTD
"""

from openpyxl import load_workbook
from pandas import DataFrame
import os

#create tables to store sig wave height values to add to dataframe and output to excel sheet
fileStore = []
WaveHeightStore = []
WavePeriodStore = []

directory = input("What is the directory where the runup files are stored? ")
directory = os.chdir(directory)
#loop through each file in the directory and opens file
numberOfFiles = 2
for file in os.listdir(directory):
    numberOfFiles += 1
    filename = os.fsdecode(file)
    if filename.endswith('.xlsm'):
        fileToCheck = filename
        #find the 1% twel for that sheet
        wb = load_workbook(filename = 'TWEL.xlsx')
        sheet_ranges = wb['sheet1']
        for i in range(2, numberOfFiles):
            valueCheck = sheet_ranges.cell(row = i, column = 1).value
            if filename == valueCheck:
                TWEL = sheet_ranges.cell(row = i, column = 3).value
                print(filename)
                print(TWEL)
            else:
                TWEL = 1
        fileToOpen = filename
        os.getcwd()
        print("file to open:", fileToOpen)
        wb = load_workbook(filename = fileToOpen, data_only = True)
        print("Loaded file")
        
                #open runup sheet and find twl with smallest difference, if the twel value is different than the 1% TWEL, find the new wave conditions
        def FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn):
            sheet_ranges = wb['Doc']
        
            #change working sheet to EventSummary sheet
            sheet_ranges = wb["EventSummary"]
            
            #search MaxTWEL values in EventSummary to find corresponding row/hour
            OriginalCheck = 100
            sheet = wb.get_sheet_by_name("EventSummary")
            for i in range(3, 152):
                valueCheck = sheet.cell(row = i, column = TWLColumn).value
                DifferenceCheck = abs(valueCheck - TWEL)
                print("Difference between", valueCheck, "and TWEL", TWEL, "is", DifferenceCheck)
                if DifferenceCheck < OriginalCheck:
                    OriginalCheck = DifferenceCheck
                    CorrectTWELRow = i
                    print("Lower difference found, changed to row", CorrectTWELRow)
                    EventNumber = sheet.cell(row = i, column = EventColumn).value
                    print("Event number changed to:", EventNumber)
                    #use MaxTWELRow and column for max wave height to find corresponding Max Wave Height
                    MaxWaveHeight = sheet.cell(row = CorrectTWELRow, column = WaveHeightColumn).value
                    
            #use event # to navigate to the correct event sheet
            EventNumberSheet = "Event" + str(EventNumber)
            sheet = wb.get_sheet_by_name(EventNumberSheet)
                
            #use MaxWaveHeight to find correct row in H0 column 
            for i in range(7, 1000):
                ValueToCheck = sheet.cell(row = i, column = EventWaveHeightColumn).value
                if ValueToCheck == MaxWaveHeight:
                    #use periodrow to find waveperiod
                    WavePeriod = sheet.cell(row = i, column = EventWavePeriodColumn).value
                    break
                else:
                    WavePeriod = "ERROR no wave period found"
                    
            #write values to excel file
            print("MaxWaveHeight =", MaxWaveHeight, "and WavePeriod =", WavePeriod)
            fileStore.append(fileToOpen)
            WaveHeightStore.append(MaxWaveHeight)
            WavePeriodStore.append(WavePeriod)
        
        if "Type 1_Stockdon_Erosion" in fileToOpen:
            print("Opened file as type 1 eroded")
            DocCell = 'C25'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 19
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7
            
            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn)
        
        elif "Type 1" in fileToOpen:
            print("opened file as type 1")
            DocCell = 'C25'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 20
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7
        
            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn)
            
        #if type 2 method cells change
        elif "Type 2" in fileToOpen:
            print("opened file as type 2")
            DocCell = 'C33'
            TWLColumn = 6
            EventColumn = 1
            WaveHeightColumn = 20
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7
        
            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn)
        
        #if type 3 method, cells and sheets change
        elif "Type 3" in fileToOpen:
            print("opened file as type 3")
            DocCell = 'C24'
            TWLColumn = 5
            EventColumn = 1
            WaveHeightColumn = 18
            EventWaveHeightColumn = 4
            EventWavePeriodColumn = 7
        
            FindSigWaveValues(DocCell, TWLColumn, EventColumn, WaveHeightColumn, EventWaveHeightColumn, EventWavePeriodColumn)
        else:
            print("ERROR")
            quit()
        
    #write values to excel file
    df = DataFrame({'File': fileStore, 'WaveHeight': WaveHeightStore, 'WavePeriod': WavePeriodStore})
    print(df)
    df.to_excel('NewValues.xlsx', sheet_name = 'sheet1', index = False)
        
        
        
        