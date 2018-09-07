# -*- coding: utf-8 -*-
"""
Created on Fri Sep  7 10:48:17 2018

@author: MANHARDTD
"""

from openpyxl import load_workbook
from pandas import DataFrame
import os

#create tables to store sig wave height values to add to dataframe and 
#output to excel sheet
file_storage = []
wave_height_storage = []
wave_period_storage = []

MAX_WAVE_ROW = 2
WAVE_HEIGHT_COLUMN = 4
WAVE_PERIOD_COLUMN = 7

directory = input("What is the directory where the JPM folders are? ")
working_directory = os.chdir(directory)
#find all the folder names in the directory, should be JPM_Analysis folder
for folder in os.listdir(working_directory):
    #add folder to end of directory and open that
    transect_folder = "%(directory)s\\%(folder)s" % {
            "directory":directory, "folder":folder}
    if os.path.isdir(transect_folder) == True:
        folder_directory = os.chdir(transect_folder)
        for file in os.listdir(transect_folder):
            filename = os.fsdecode(file)
            if filename.endswith('5Cases.xlsx') and "~$" not in filename:
                file_to_open = filename
                os.getcwd()
                print(file_to_open)
                wb = load_workbook(filename = file_to_open, data_only=True)        
                sheet = wb["Sheet1"]        
                wave_height = sheet.cell(row=MAX_WAVE_ROW, column=WAVE_HEIGHT_COLUMN).value
                wave_height_storage.append(wave_height)
                wave_period = sheet.cell(row=MAX_WAVE_ROW, column=WAVE_PERIOD_COLUMN).value
                wave_period_storage.append(wave_period)        
    #create dataframe with file_storage, wave_height, and wave_period
                file_storage.append(file_to_open)
        df = DataFrame({'File': file_storage, 
                        'Wave Height': wave_height_storage, 
                        'Wave Period': wave_period_storage})
        print(df)
        working_directory = os.chdir(directory)
        df.to_excel('FISTable17WaveConditions.xlsx', sheet_name = 'sheet1', index = False)
 