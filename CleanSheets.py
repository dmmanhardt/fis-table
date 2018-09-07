# -*- coding: utf-8 -*-
"""
Created on Fri Jul  6 13:51:05 2018

@author: MANHARDTD
"""
from openpyxl import load_workbook
import os

def CleanSheet(ColumnToCheck, RunupColumn, HmoColumn):
    sheet = wb.get_sheet_by_name("EventSummary")
    RowNumber = 3
    ChangesMade = False
    for EventNumber in range(1, 150):
        CellToCheck = sheet.cell(row = RowNumber, column = ColumnToCheck).value
        print(CellToCheck)
        RowNumber += 1
        if CellToCheck == "#VALUE!":
            print("Found #Value! in event ", EventNumber)
            EventNumberSheet = "Event" + str(EventNumber)
            sheet = wb.get_sheet_by_name(EventNumberSheet)
            # delete runup
            sheet.cell(row = 7, column = RunupColumn).value = None
            print("Deleted runup cell for event number ", EventNumber)
            # delete hmo
            if HmoColumn == "ERROR":
                break
            else:
                sheet.cell(row = 7, column = HmoColumn).value = None
                print("Deleted Hmo cell for event number ", EventNumber)
                ChangesMade = True
    if ChangesMade == True:
        wb.save(fileToOpen)

directory = input("What is the directory where the runup files are stored? ")
directory = os.chdir(directory)
#loop through each file in the directory and opens file
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith('.xlsm'):
        fileToOpen = filename
        os.getcwd()
        print(fileToOpen)
        wb = load_workbook(fileToOpen, data_only = False, read_only = False, keep_vba = True)
        
        if "Type 1_Stockdon_Erosion" in fileToOpen:
            print("Opened file as type 1 eroded")
            ColumnToCheck = 19
            RunupColumn = 22
            HmoColumn = "ERROR"
            CleanSheet(ColumnToCheck, RunupColumn, HmoColumn)

        
        elif "Type 1" in fileToOpen:
            print("opened file as type 1")
            ColumnToCheck = 20
            RunupColumn = 22
            HmoColumn = "ERROR"
            CleanSheet(ColumnToCheck, RunupColumn, HmoColumn)

            
        #if type 2 method cells change
        elif "Type 2" in fileToOpen:
            print("opened file as type 2")
            ColumnToCheck = 22
            RunupColumn = 45
            HmoColumn = 35
            CleanSheet(ColumnToCheck, RunupColumn, HmoColumn)

        
        #if type 3 method, cells and sheets change
        elif "Type 3" in fileToOpen:
            print("opened file as type 3")
            ColumnToCheck = 19
            RunupColumn = 39
            HmoColumn = 30
            CleanSheet(ColumnToCheck, RunupColumn, HmoColumn)

        else:
            print("ERROR")
            quit()