# -*- coding: utf-8 -*-
"""
Created on Fri Aug 24 10:47:13 2018

@author: MANHARDTD
"""

# Finds the max wave height associated with the maximum runup for each
# engineering summary sheet. Outputs info as excel file with transect no.
# and wave wave height

from openpyxl import load_workbook
from pandas import DataFrame
import os

#create tables to store sig wave height values to add to dataframe and output to excel sheet
file_storage = []
max_wave_height = []
runup_wave_height = []

def find_max_wave_height(max_Hs_column, runup_column, event_column, wave_height_column, event_runup_column):
    sheet = wb["EventSummary"]
    #find max Hs
    max_Hs = 0
    for i in range(3, 152):
        max_Hs_check = sheet.cell(row = i, column = max_Hs_column).value
        try:
            if max_Hs_check > max_Hs:
                max_Hs = max_Hs_check
        except:
            print("error for %(row)i" % {"row":i})
    max_runup = 0
    found_bigger_runup = False
    #find max runup
    for i in range(3, 152):
        value_check = float(sheet.cell(row = i, column = runup_column).value)
        print("value_check = %(check)i" % {"check":value_check})
        if value_check > max_runup:
            max_runup = value_check
            max_runup_row = i
            found_bigger_runup = True
            event_number = sheet.cell(row = max_runup_row, column = event_column).value           
    print("found_bigger_runup = ", found_bigger_runup)
    if found_bigger_runup == True:
        #use event # to navigate to the correct event sheet
        event_numberSheet = "Event" + str(event_number)
        sheet = wb[event_numberSheet]
            
        #use max_runup to find correct row in max_wave_height_column 
        for i in range(7, 1000):
            runup_check = sheet.cell(row = i, column = event_runup_column).value
            if runup_check == max_runup:
                #use periodrow to find waveperiod
                wave_height = sheet.cell(row = i, column = wave_height_column).value
                break
            else:
                wave_height = "ERROR no wave height found"
                
            #write values to excel file
        file_storage.append(file_to_open)
        max_wave_height.append(max_Hs)
        runup_wave_height.append(wave_height)


#Path where excel files are located, needs to be updated for each county
    
directory = input("What is the directory where the runup files are stored? ")
directory = os.chdir(directory)
#finds all runup files in the pointed directory, loops over them and opens them
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith('.xlsm') and "~$" not in filename:
        file_to_open = filename
        os.getcwd()
        print(file_to_open)
        wb = load_workbook(filename = file_to_open, data_only=True)

#check what method (type 1-3) and update parameters for that method, then run find_max_wave_height()

        if "Type 1_Stockdon_Erosion" in file_to_open:
            print("Opened file as type 1 eroded")
            max_Hs_column = 19
            runup_column = 5
            event_column = 1
            event_runup_column = 22
            wave_height_column = 4
            
            find_max_wave_height(max_Hs_column, runup_column, event_column, wave_height_column, event_runup_column)

        elif "Type 1" in file_to_open:
            print("opened file as type 1")
            max_Hs_column = 20
            runup_column = 5
            event_column = 1
            event_runup_column = 22
            wave_height_column = 4

            find_max_wave_height(max_Hs_column, runup_column, event_column, wave_height_column, event_runup_column)
            
        elif "Type 2" in file_to_open:
            print("opened file as type 2")
            max_Hs_column = 20
            runup_column = 5
            event_column = 1
            event_runup_column = 45
            wave_height_column = 4

            find_max_wave_height(max_Hs_column, runup_column, event_column, wave_height_column, event_runup_column)
        
        elif "Type 3" in file_to_open:
            print("opened file as type 3")
            max_Hs_column = 18
            runup_column = 4
            event_column = 1
            event_runup_column = 40
            wave_height_column = 4

            find_max_wave_height(max_Hs_column, runup_column, event_column, wave_height_column, event_runup_column)
        else:
            print("ERROR")
            quit()
    #write values to excel file
    df = DataFrame({'File': file_storage, 'Max Hs': max_wave_height, 'Runup Wave Height': runup_wave_height})
    print(df)
    df.to_excel('MaxWaveHeights.xlsx', sheet_name = 'sheet1', index = False)